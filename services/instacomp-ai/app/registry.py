from __future__ import annotations

import csv
import hashlib
import json
import os
import shutil
import sqlite3
import tempfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook
from pydantic import ValidationError

from .checklist_schema import CHECKLIST_COLUMNS, ChecklistRow
from .models import CardIdentity, ChecklistOutcome, ChecklistResult


def _bool(value: object) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y"}


def _int_or_none(value: object) -> int | None:
    text = str(value or "").strip()
    return int(text) if text else None


def _normalize(value: object) -> str:
    return " ".join(str(value or "").casefold().split())


def _card_number(value: object) -> str:
    return _normalize(value).replace(" ", "").lstrip("#")


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _row_from_mapping(mapping: dict[str, object]) -> ChecklistRow:
    payload = {column: mapping.get(column) for column in CHECKLIST_COLUMNS}
    payload["rookie"] = _bool(payload.get("rookie"))
    payload["autograph"] = _bool(payload.get("autograph"))
    payload["memorabilia"] = _bool(payload.get("memorabilia"))
    payload["serial_run"] = _int_or_none(payload.get("serial_run"))
    return ChecklistRow.model_validate(payload)


def read_rows(path: Path) -> Iterable[ChecklistRow]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            for mapping in csv.DictReader(handle):
                yield _row_from_mapping(dict(mapping))
        return
    if suffix == ".json":
        payload = json.loads(path.read_text(encoding="utf-8"))
        records = payload if isinstance(payload, list) else payload.get("rows", [])
        for mapping in records:
            if isinstance(mapping, dict):
                yield _row_from_mapping(mapping)
        return
    if suffix in {".xlsx", ".xlsm"}:
        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            sheet = workbook.active
            values = sheet.iter_rows(values_only=True)
            headers = [str(value or "").strip() for value in next(values)]
            for values_row in values:
                mapping = dict(zip(headers, values_row))
                if any(value not in (None, "") for value in values_row):
                    yield _row_from_mapping(mapping)
        finally:
            workbook.close()
        return
    raise ValueError(f"Unsupported registry source: {path.name}")


class RegistryBuilder:
    def __init__(self, mirror_root: Path, registry_path: Path, quarantine_root: Path):
        self.mirror_root = mirror_root
        self.registry_path = registry_path
        self.quarantine_root = quarantine_root

    def build(self) -> dict[str, object]:
        self.registry_path.parent.mkdir(parents=True, exist_ok=True)
        self.quarantine_root.mkdir(parents=True, exist_ok=True)
        imported_files = 0
        imported_rows = 0
        rejected_files: list[dict[str, object]] = []
        duplicate_rows = 0
        previous_registry_existed = self.registry_path.exists()
        activated = False

        with tempfile.TemporaryDirectory(prefix="instacomp-registry-") as temporary:
            candidate = Path(temporary) / "registry.sqlite3"
            connection = sqlite3.connect(candidate)
            try:
                self._create_schema(connection)
                for source in sorted(self.mirror_root.rglob("*")):
                    if not source.is_file() or source.suffix.lower() not in {
                        ".csv",
                        ".json",
                        ".xlsx",
                        ".xlsm",
                    }:
                        continue
                    try:
                        rows = list(read_rows(source))
                        if not rows:
                            raise ValueError("No checklist rows found")
                        with connection:
                            for row in rows:
                                try:
                                    self._insert(connection, row, source)
                                    imported_rows += 1
                                except sqlite3.IntegrityError:
                                    duplicate_rows += 1
                        imported_files += 1
                    except (
                        OSError,
                        ValueError,
                        ValidationError,
                        json.JSONDecodeError,
                        StopIteration,
                        TypeError,
                    ) as exc:
                        quarantine = self._quarantine(source, exc)
                        rejected_files.append(
                            {
                                "file": str(source),
                                "error": str(exc),
                                "quarantine_file": str(quarantine["file"]),
                                "quarantine_receipt": str(quarantine["receipt"]),
                            }
                        )
                connection.execute("PRAGMA optimize")
                integrity = connection.execute("PRAGMA integrity_check").fetchone()[0]
                if integrity != "ok":
                    raise RuntimeError(f"Registry integrity check failed: {integrity}")
                connection.commit()
            finally:
                connection.close()

            candidate_clean = imported_rows > 0 and not rejected_files
            if candidate_clean:
                os.replace(candidate, self.registry_path)
                activated = True

        active_rows = self._active_row_count()
        receipt = {
            "schema": "tcos.instacomp-ai.registry-build.v2",
            "created_at": datetime.now(timezone.utc).isoformat(),
            "registry_path": str(self.registry_path),
            "candidate_imported_files": imported_files,
            "candidate_imported_rows": imported_rows,
            "duplicate_rows": duplicate_rows,
            "rejected_files": rejected_files,
            "activated": activated,
            "previous_registry_retained": bool(
                previous_registry_existed and not activated
            ),
            "active_rows": active_rows,
            "ready": active_rows > 0,
        }
        self._atomic_json(self.registry_path.parent / "latest-build.json", receipt)
        return receipt

    def _quarantine(self, source: Path, error: Exception) -> dict[str, Path]:
        stamp = datetime.now(timezone.utc).strftime("%Y%m%dT%H%M%S.%fZ")
        try:
            relative = source.relative_to(self.mirror_root)
        except ValueError:
            relative = Path(source.name)
        destination = self.quarantine_root / stamp / relative
        destination.parent.mkdir(parents=True, exist_ok=True)
        shutil.copy2(source, destination)
        receipt = destination.with_suffix(destination.suffix + ".error.json")
        self._atomic_json(
            receipt,
            {
                "schema": "tcos.instacomp-ai.checklist-quarantine.v1",
                "created_at": datetime.now(timezone.utc).isoformat(),
                "source": str(source),
                "quarantine_file": str(destination),
                "source_sha256": _sha256(source),
                "error_type": type(error).__name__,
                "error": str(error),
            },
        )
        return {"file": destination, "receipt": receipt}

    def _active_row_count(self) -> int:
        if not self.registry_path.exists():
            return 0
        try:
            with sqlite3.connect(
                f"file:{self.registry_path}?mode=ro", uri=True
            ) as connection:
                return int(
                    connection.execute("SELECT COUNT(*) FROM checklist_cards").fetchone()[0]
                )
        except sqlite3.Error:
            return 0

    @staticmethod
    def _atomic_json(path: Path, payload: dict[str, object]) -> None:
        temporary = path.with_suffix(path.suffix + ".partial")
        temporary.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        os.replace(temporary, path)

    @staticmethod
    def _create_schema(connection: sqlite3.Connection) -> None:
        connection.executescript(
            """
            PRAGMA journal_mode=DELETE;
            PRAGMA synchronous=FULL;
            CREATE TABLE checklist_cards (
              identity_id TEXT PRIMARY KEY,
              source_file TEXT NOT NULL,
              source_name TEXT NOT NULL,
              source_release_id TEXT NOT NULL,
              source_version TEXT NOT NULL,
              source_receipt TEXT NOT NULL,
              sport TEXT NOT NULL,
              league TEXT,
              year TEXT NOT NULL,
              manufacturer TEXT,
              brand TEXT NOT NULL,
              set_name TEXT NOT NULL,
              subset TEXT,
              player TEXT NOT NULL,
              team TEXT,
              card_number TEXT NOT NULL,
              parallel TEXT NOT NULL,
              variation TEXT,
              serial_run INTEGER,
              rookie INTEGER NOT NULL,
              autograph INTEGER NOT NULL,
              memorabilia INTEGER NOT NULL,
              language_code TEXT NOT NULL,
              notes TEXT,
              row_json TEXT NOT NULL
            );
            CREATE INDEX checklist_lookup ON checklist_cards(year, set_name, player, card_number);
            CREATE INDEX checklist_set_lookup ON checklist_cards(year, set_name);
            """
        )

    @staticmethod
    def _insert(connection: sqlite3.Connection, row: ChecklistRow, source: Path) -> None:
        values = row.model_dump(mode="json")
        connection.execute(
            """INSERT INTO checklist_cards VALUES (
            :identity_id,:source_file,:source_name,:source_release_id,:source_version,:source_receipt,
            :sport,:league,:year,:manufacturer,:brand,:set_name,:subset,:player,:team,:card_number,
            :parallel,:variation,:serial_run,:rookie,:autograph,:memorabilia,:language_code,:notes,:row_json)
            """,
            {
                **values,
                "identity_id": row.identity_fingerprint(),
                "source_file": str(source),
                "rookie": int(row.rookie),
                "autograph": int(row.autograph),
                "memorabilia": int(row.memorabilia),
                "row_json": json.dumps(values, sort_keys=True),
            },
        )


class SQLiteChecklistRegistry:
    def __init__(self, registry_path: Path):
        self.registry_path = registry_path

    async def health(self) -> bool:
        if not self.registry_path.exists():
            return False
        try:
            with sqlite3.connect(f"file:{self.registry_path}?mode=ro", uri=True) as connection:
                return connection.execute("SELECT COUNT(*) FROM checklist_cards").fetchone()[0] > 0
        except sqlite3.Error:
            return False

    async def match(self, identity: CardIdentity) -> ChecklistResult:
        required = {
            "year": identity.year,
            "set_name": identity.set_name,
            "player": identity.player,
            "card_number": identity.card_number,
        }
        missing = [key for key, value in required.items() if not value]
        if missing:
            return ChecklistResult(
                outcome=ChecklistOutcome.INPUT_INCOMPLETE,
                reasons=[f"Missing identity fields: {', '.join(missing)}"],
            )
        if not await self.health():
            return ChecklistResult(
                outcome=ChecklistOutcome.NOT_CONFIGURED,
                reasons=["No active checklist registry has been built yet."],
            )

        with sqlite3.connect(f"file:{self.registry_path}?mode=ro", uri=True) as connection:
            connection.row_factory = sqlite3.Row
            rows = connection.execute(
                "SELECT * FROM checklist_cards WHERE lower(year)=? AND lower(set_name)=?",
                (_normalize(identity.year), _normalize(identity.set_name)),
            ).fetchall()
        if not rows:
            return ChecklistResult(
                outcome=ChecklistOutcome.SET_ABSENT,
                reasons=["The requested year and set are not present in the active registry."],
            )

        exact = []
        for row in rows:
            if _normalize(row["player"]) != _normalize(identity.player):
                continue
            if _card_number(row["card_number"]) != _card_number(identity.card_number):
                continue
            requested_parallel = _normalize(identity.parallel or "Base")
            requested_variation = _normalize(identity.variation)
            if requested_parallel and _normalize(row["parallel"] or "Base") != requested_parallel:
                continue
            if requested_variation and _normalize(row["variation"]) != requested_variation:
                continue
            exact.append(row)

        if len(exact) != 1:
            return ChecklistResult(
                outcome=ChecklistOutcome.SET_PRESENT_NO_EXACT_MATCH,
                candidate_count=len(exact) if exact else len(rows),
                reasons=[
                    "The set exists, but the available identity fields do not resolve to exactly one checklist row."
                ],
            )

        row = exact[0]
        card = CardIdentity(
            sport=row["sport"],
            league=row["league"],
            year=row["year"],
            manufacturer=row["manufacturer"],
            brand=row["brand"],
            set_name=row["set_name"],
            subset=row["subset"],
            player=row["player"],
            team=row["team"],
            card_number=row["card_number"],
            parallel=row["parallel"],
            variation=row["variation"],
            serial_run=row["serial_run"],
            rookie=bool(row["rookie"]),
            autograph=bool(row["autograph"]),
            memorabilia=bool(row["memorabilia"]),
        )
        return ChecklistResult(
            outcome=ChecklistOutcome.EXACT_MATCH,
            identity_id=row["identity_id"],
            identity=card,
            candidate_count=1,
            reasons=["One exact row matched the active, versioned checklist registry."],
            source_receipts=[row["source_receipt"]],
        )
